from __future__ import annotations

import csv
import io
import math
import hashlib
import json
import logging
import os
import re
import secrets
import signal
import tempfile
import time
import unicodedata
import threading
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from flask import Flask, jsonify, request, send_file, send_from_directory, session
from openpyxl.cell import WriteOnlyCell
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from rapidfuzz import fuzz
except Exception:  # pragma: no cover
    fuzz = None

try:
    import xlrd
except Exception:  # pragma: no cover
    xlrd = None

BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"

app = Flask(__name__, static_folder=str(STATIC_DIR), static_url_path="/static")
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024
app.config["SECRET_KEY"] = os.environ.get(
    "GRAMBOOK_SECRET_KEY", "grambook-development-secret-key"
)

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger("grambook")

MAX_ROWS = 50_000
PREVIEW_ROWS = 10
HEADER_SEARCH_ROWS = 10
HEADER_SEARCH_SPAN = 5
# FIXED: unified relaxed threshold for noisy real-world government headers.
FUZZY_MATCH_THRESHOLD = 58
MAPPING_MIN_RATIO = 0.50
MAPPING_MIN_WEIGHTED_CONFIDENCE = 60.0
MAPPING_CORE_CONFIDENCE = 75.0
KEY_NON_EMPTY_HARD_MIN = 0.60
KEY_UNIQUE_HARD_MIN = 0.70
KEY_BORDERLINE_MIN = 0.80
NUMERIC_TOLERANCE = Decimal("0.005")

CSRF_TOKEN = hashlib.sha256(
    f"{app.config['SECRET_KEY']}::csrf".encode("utf-8")
).hexdigest()
RESULT_CACHE: dict[str, dict[str, Any]] = {}
RESULT_CACHE_ORDER: list[str] = []
RESULT_CACHE_LIMIT = 16
RESULT_CACHE_LOCK = threading.RLock()
CACHE_SESSION_KEY = "grambook_session_id"
DEFAULT_PAGE_SIZE = 200
MAX_PAGE_SIZE = 500
REQUEST_TIMEOUT_SECONDS = 30
JSON_ROW_LIMIT = 1000

COLUMN_MISSING = "__COLUMN_MISSING__"
VALUE_MISSING = "__VALUE_MISSING__"

VALID_DATE_MODES = {"auto", "strict", "day_first", "month_first"}
FIXED_ADMIN_HEADER_ROW = 2
FIXED_ADMIN_HEADER_SPAN = 3
FIXED_SUV_HEADER_ROW = 1
FIXED_SUV_HEADER_SPAN = 2

ZERO_WIDTH_RE = re.compile(r"[\u200B-\u200D\uFEFF]")
CONTROL_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")
SPACE_RE = re.compile(r"\s+")
COMPACT_RE = re.compile(r"[^\w]+", re.UNICODE)
NUMERIC_RE = re.compile(r"^[+-]?\d+(?:\.\d+)?$")
BOOLEAN_TRUE_TEXTS = {"true", "yes", "y", "on", "1"}
BOOLEAN_FALSE_TEXTS = {"false", "no", "n", "off", "0"}

DIGIT_TRANSLATION = str.maketrans(
    {
        "\u0966": "0",
        "\u0967": "1",
        "\u0968": "2",
        "\u0969": "3",
        "\u096a": "4",
        "\u096b": "5",
        "\u096c": "6",
        "\u096d": "7",
        "\u096e": "8",
        "\u096f": "9",
        "\u0ae6": "0",
        "\u0ae7": "1",
        "\u0ae8": "2",
        "\u0ae9": "3",
        "\u0aea": "4",
        "\u0aeb": "5",
        "\u0aec": "6",
        "\u0aed": "7",
        "\u0aee": "8",
        "\u0aef": "9",
    }
)

HEADER_GROUPS: dict[str, list[str]] = {
    "Identifier": [
        "id",
        "code",
        "ref",
        "reference",
        "number",
        "no",
        "num",
        "key",
        "uid",
        "identifier",
        "ક્રમ",
        "ક્રમાંક",
        "નં",
        "નંબર",
        "કોડ",
        "ઓળખ",
    ],
    "Name": ["name", "નામ", "beneficiary", "holder", "citizen", "person"],
    "Amount": ["amount", "amt", "value", "રકમ", "ફી", "tax", "payment"],
    "Date": ["date", "તારીખ", "દિન", "day", "dob"],
    "Mobile": ["mobile", "phone", "contact", "મોબાઇલ", "whatsapp"],
    "Address": ["address", "addr", "સરનામું", "સરનામો", "location"],
    "Location": ["village", "ward", "taluka", "district", "ગામ", "વોર્ડ", "જિલ્લો"],
    "Status": ["status", "સ્થિતિ", "state", "remarks", "remark", "note", "notes"],
}


class ReconciliationError(Exception):
    pass


@dataclass
class ParsedDataset:
    rows: list[dict[str, str]]
    normalized_rows: list[dict[str, str]]
    columns: list[str]
    column_meta: list[dict[str, Any]]
    header_row_index: int
    header_row_span: int
    parser_notes: list[str] = field(default_factory=list)
    source_format: str = "unknown"
    row_numbers: list[int] = field(default_factory=list)


@dataclass
class RowRecord:
    index: int
    row_number: int
    key_display: str
    key_norm: str
    row: dict[str, str]
    norm_row: dict[str, str]
    fingerprint: str
    key_missing: bool


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float):
        if not math.isfinite(value):
            return ""
        if value.is_integer():
            return str(int(value))
        text = f"{value:.15f}".rstrip("0").rstrip(".")
        return text or "0"
    text = str(value)
    text = CONTROL_RE.sub("", text)
    text = ZERO_WIDTH_RE.sub("", text)
    text = unicodedata.normalize("NFKC", text)
    text = text.translate(DIGIT_TRANSLATION)
    return text.strip()


def _normalize_gujarati_terms(text: str) -> str:
    # FIXED: normalize common Gujarati variants before compact matching.
    out = text
    out = unicodedata.normalize("NFKC", out)
    out = out.replace("\u00a0", " ")
    out = out.replace("_", " ")
    out = out.replace("ـ", " ")
    out = re.sub(r"[\u200B-\u200D\uFEFF]", "", out)
    out = re.sub(r"ઘર\s*વેરો", "ઘરવેરા", out)
    out = re.sub(r"વેરો\b", "વેરા", out)
    out = re.sub(r"વેરા\s*gs\b", "વેરા gs", out, flags=re.IGNORECASE)
    out = re.sub(r"\b(\w+)\s*[_-](\d+)\b", r"\1 \2", out)
    out = re.sub(r"\s*\(\s*(\d+)\s*\)\s*$", r" \1", out)
    out = re.sub(r"\s+", " ", out).strip()
    return out


def _normalize_for_compare(value: Any) -> str:
    text = _clean_text(value)
    if not text:
        return ""
    text = _normalize_gujarati_terms(text)
    return SPACE_RE.sub(" ", text).casefold().strip()


def _normalize_for_match(value: Any) -> str:
    text = _clean_text(value)
    if not text:
        return ""
    text = _normalize_gujarati_terms(text)
    text = COMPACT_RE.sub(" ", text)
    return SPACE_RE.sub(" ", text).casefold().strip()


def _normalize_header_compact(value: Any) -> str:
    text = _normalize_for_match(value)
    text = re.sub(r"[_\-\s]*\d+$", "", text)
    return text.replace(" ", "")


def _tokenize_for_match(value: Any) -> list[str]:
    text = _normalize_for_match(value)
    if not text:
        return []
    return [token for token in text.split() if token]


def _numeric_text(value: Any) -> Decimal | None:
    text = _clean_text(value)
    if not text:
        return None
    text = text.replace(",", "").replace(" ", "")
    if not NUMERIC_RE.fullmatch(text):
        return None
    try:
        num = Decimal(text)
    except InvalidOperation:
        return None
    if not num.is_finite():
        return None
    return num


def _boolean_text(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    text = _normalize_for_compare(value)
    if text in BOOLEAN_TRUE_TEXTS:
        return True
    if text in BOOLEAN_FALSE_TEXTS:
        return False
    return None


def _date_text(value: Any, mode: str = "auto") -> str | None:
    text = _clean_text(value)
    if not text:
        return None
    if isinstance(value, (date, datetime)):
        return (
            value.date().isoformat()
            if isinstance(value, datetime)
            else value.isoformat()
        )
    candidate = _normalize_for_compare(text)
    candidate = candidate.replace(".", "/").replace("-", "/")
    if not re.fullmatch(r"\d{1,4}/\d{1,2}/\d{1,4}", candidate):
        return None
    parts = [p for p in candidate.split("/") if p]
    if len(parts) != 3:
        return None
    if len(parts[0]) == 4:
        try:
            dt = datetime.strptime(candidate, "%Y/%m/%d")
            return dt.date().isoformat()
        except ValueError:
            return None

    day, month, year = parts
    if len(day) <= 2 and len(month) <= 2:
        day_i = int(day)
        month_i = int(month)
        ambiguous = day_i <= 12 and month_i <= 12
        if mode == "strict" and ambiguous:
            raise ReconciliationError(
                f"Ambiguous date '{text}' requires a date mode selection."
            )
        if mode == "day_first":
            fmts = ("%d/%m/%Y", "%d/%m/%y")
        elif mode == "month_first":
            fmts = ("%m/%d/%Y", "%m/%d/%y")
        else:
            if day_i > 12 and month_i <= 12:
                fmts = ("%d/%m/%Y", "%d/%m/%y")
            elif month_i > 12 and day_i <= 12:
                fmts = ("%m/%d/%Y", "%m/%d/%y")
            else:
                return None
        for fmt in fmts:
            try:
                dt = datetime.strptime(candidate, fmt)
                return dt.date().isoformat()
            except ValueError:
                continue
    return None


def _values_equal(left: Any, right: Any, date_mode: str = "auto") -> bool:
    left_text = _normalize_for_compare(left)
    right_text = _normalize_for_compare(right)
    if left_text == "" and right_text == "":
        return True
    if (left_text == "") != (right_text == ""):
        return False
    left_bool = _boolean_text(left)
    right_bool = _boolean_text(right)
    if left_bool is not None and right_bool is not None:
        left_is_explicit_bool = (
            isinstance(left, bool)
            or left_text in BOOLEAN_TRUE_TEXTS
            or left_text in BOOLEAN_FALSE_TEXTS
        )
        right_is_explicit_bool = (
            isinstance(right, bool)
            or right_text in BOOLEAN_TRUE_TEXTS
            or right_text in BOOLEAN_FALSE_TEXTS
        )
        if left_is_explicit_bool or right_is_explicit_bool:
            return left_bool == right_bool
    left_date = _date_text(left, date_mode)
    right_date = _date_text(right, date_mode)
    if left_date is not None and right_date is not None:
        return left_date == right_date
    left_num = _numeric_text(left)
    right_num = _numeric_text(right)
    if left_num is not None and right_num is not None:
        # FIXED: numeric comparison uses a tighter tolerance to avoid false positives.
        return abs(left_num - right_num) < NUMERIC_TOLERANCE
    if (left_num is None) != (right_num is None):
        return False
    return left_text == right_text


def _fingerprint_row(row: dict[str, str], column_order: list[str]) -> str:
    return " | ".join(_normalize_for_compare(row.get(col, "")) for col in column_order)


def _guess_group(name: str) -> str:
    norm = _normalize_header_compact(name)
    for group, aliases in HEADER_GROUPS.items():
        for alias in aliases:
            if _normalize_header_compact(alias) in norm:
                return group
    return "Other"


def _column_meta(columns: list[str]) -> list[dict[str, Any]]:
    meta: list[dict[str, Any]] = []
    for idx, col in enumerate(columns):
        meta.append(
            {
                "column": col,
                "normalized": _normalize_header_compact(col),
                "group": _guess_group(col),
                "hierarchy": _normalize_for_compare(col),
                "index": idx,
            }
        )
    return meta


def _file_bytes(upload) -> tuple[bytes, str]:
    if upload is None:
        raise ReconciliationError("Both files are required.")
    try:
        upload.stream.seek(0)
    except Exception:
        pass
    data = upload.read()
    try:
        upload.stream.seek(0)
    except Exception:
        pass
    if not data:
        raise ReconciliationError(
            f"Uploaded file '{upload.filename or 'file'}' is empty."
        )
    return data, (upload.filename or "").lower()


def _decode_csv_bytes(file_bytes: bytes) -> io.StringIO:
    for encoding in ("utf-8-sig", "utf-8", "utf-16", "cp1252", "latin1"):
        try:
            return io.StringIO(file_bytes.decode(encoding))
        except UnicodeDecodeError:
            continue
    raise ReconciliationError("CSV encoding is unsupported or the file is corrupt.")


def _normalize_matrix_width(matrix: list[list[str]]) -> list[list[str]]:
    if not matrix:
        return matrix
    width = max((len(row) for row in matrix), default=0)
    out: list[list[str]] = []
    for row in matrix:
        row = list(row)
        if len(row) < width:
            row.extend([""] * (width - len(row)))
        else:
            row = row[:width]
        out.append(row)
    return out


def _prune_blank_rows(matrix: list[list[str]]) -> tuple[list[list[str]], list[int]]:
    kept: list[list[str]] = []
    row_numbers: list[int] = []
    for idx, row in enumerate(matrix, start=1):
        if any(_normalize_for_compare(cell) for cell in row):
            kept.append(row)
            row_numbers.append(idx)
    return kept, row_numbers


def _parse_csv_matrix(file_bytes: bytes) -> tuple[list[list[str]], list[int]]:
    sio = _decode_csv_bytes(file_bytes)
    sample = sio.read(4096)
    sio.seek(0)
    dialect = csv.excel
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        pass
    matrix = [[_clean_text(cell) for cell in row] for row in csv.reader(sio, dialect)]
    matrix = _normalize_matrix_width(matrix)
    return _prune_blank_rows(matrix)


def _parse_xls_matrix(file_bytes: bytes) -> tuple[list[list[str]], list[int]]:
    if xlrd is None:
        raise ReconciliationError(".xls support requires xlrd==2.0.1")
    try:
        wb = xlrd.open_workbook(file_contents=file_bytes)
        sh = wb.sheet_by_index(0)
    except Exception as exc:
        raise ReconciliationError(f"Unable to read .xls workbook: {exc}") from exc
    matrix = [
        [_clean_text(sh.cell_value(r, c)) for c in range(sh.ncols)]
        for r in range(sh.nrows)
    ]
    matrix = _normalize_matrix_width(matrix)
    return _prune_blank_rows(matrix)


def _parse_xlsx_matrix(file_bytes: bytes) -> tuple[list[list[str]], list[int]]:
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
        ws = wb.worksheets[0]
    except Exception as exc:
        raise ReconciliationError(f"Unable to read .xlsx workbook: {exc}") from exc
    merged_map: dict[str, str] = {}
    merged_cells = getattr(ws, "merged_cells", None)
    if merged_cells is not None:
        for merged_range in merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left = ws.cell(min_row, min_col).value
            if top_left is None:
                continue
            cleaned = _clean_text(top_left)
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    merged_map[f"{row}:{col}"] = cleaned
    matrix: list[list[str]] = []
    row_numbers: list[int] = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        display_row: list[str] = []
        for col_idx, cell in enumerate(row, start=1):
            if cell is None:
                cell = merged_map.get(f"{row_idx}:{col_idx}", "")
            display_row.append(_clean_text(cell))
        if any(_normalize_for_compare(cell) for cell in display_row):
            matrix.append(display_row)
            row_numbers.append(row_idx)
    matrix = _normalize_matrix_width(matrix)
    wb.close()
    return matrix, row_numbers


def _parse_matrix_from_upload(upload) -> tuple[list[list[str]], list[int], str]:
    file_bytes, filename = _file_bytes(upload)
    if filename.endswith(".csv"):
        matrix, rows = _parse_csv_matrix(file_bytes)
        return matrix, rows, "csv"
    if filename.endswith(".xls"):
        matrix, rows = _parse_xls_matrix(file_bytes)
        return matrix, rows, "xls"
    if filename.endswith(".xlsx"):
        matrix, rows = _parse_xlsx_matrix(file_bytes)
        return matrix, rows, "xlsx"
    raise ReconciliationError("Unsupported file format. Upload .csv, .xls, or .xlsx")


def _pad_header_rows(matrix: list[list[str]], start: int, span: int) -> int:
    if start < 0:
        start = 0
    if start >= len(matrix):
        return 0
    if span < 1:
        span = 1
    return min(span, len(matrix) - start)


def _build_header_names(matrix: list[list[str]], start: int, span: int) -> list[str]:
    width = max((len(row) for row in matrix), default=0)
    names: list[str] = []
    seen: dict[str, int] = {}
    for c in range(width):
        parts: list[str] = []
        for r in range(start, min(len(matrix), start + span)):
            if c < len(matrix[r]):
                cell = _clean_text(matrix[r][c])
                if cell:
                    parts.append(cell)
        raw = SPACE_RE.sub(" ", " ".join(parts).strip())
        if not raw:
            raw = f"Column {c + 1}"
        base = raw
        seen_key = _normalize_for_compare(base) or base.casefold()
        if seen_key in seen:
            seen[seen_key] += 1
            raw = f"{base}_{seen[seen_key]}"
        else:
            seen[seen_key] = 1
        names.append(raw)
    return names


def _header_score(start: int, span: int, names: list[str]) -> float:
    if not names:
        return -1.0
    non_blank = 0
    unique_norm = set()
    alpha_cells = 0
    token_hits = 0
    numeric_penalty = 0.0
    for name in names:
        norm = _normalize_header_compact(name)
        if norm and not norm.startswith("column"):
            non_blank += 1
            unique_norm.add(norm)
            tokens = _tokenize_for_match(name)
            digit_count = sum(ch.isdigit() for ch in name)
            letter_count = sum(ch.isalpha() for ch in name)
            if digit_count and digit_count >= letter_count:
                numeric_penalty += 1.5
            if re.search(r"[^\W\d_]", name, flags=re.UNICODE):
                alpha_cells += 1
            if any(
                token in norm
                for token in ("gs", "id", "નં", "નંબ", "કોડ", "ક્રમ", "સરનામું", "વેરા")
            ):
                alpha_cells += 1
            if any(
                token
                in {
                    "gs",
                    "id",
                    "no",
                    "num",
                    "code",
                    "ref",
                    "key",
                    "identifier",
                    "નં",
                    "નંબર",
                    "ક્રમ",
                    "કોડ",
                }
                for token in tokens
            ):
                token_hits += 1
    duplicate_penalty = len(names) - len(unique_norm)
    score = (non_blank * 4.0) + (len(unique_norm) * 2.5) + (alpha_cells * 1.5)
    score -= numeric_penalty * 2.0
    # FIXED: stable bonuses for common government header terms.
    score += sum(
        3.0
        for name in names
        if any(
            token in _normalize_for_match(name)
            for token in ("gs", "id", "નં", "નંબર", "કોડ", "ક્રમ", "વેરા")
        )
    )
    score += token_hits * 2.0
    score -= duplicate_penalty * 2.0
    score -= start * 0.25
    score += span * 0.75
    if non_blank < max(1, len(names) // 2):
        score -= 5.0
    return score


def _detect_header_spec(matrix: list[list[str]]) -> tuple[int, int, list[str]]:
    if not matrix:
        raise ReconciliationError("No usable rows found in file.")
    best_start = 0
    best_span = 1
    best_names = _build_header_names(matrix, 0, 1)
    best_score = -(10**9)
    search_rows = min(HEADER_SEARCH_ROWS, len(matrix))
    for start in range(search_rows):
        max_span = min(HEADER_SEARCH_SPAN, len(matrix) - start)
        for span in range(1, max_span + 1):
            names = _build_header_names(matrix, start, span)
            score = _header_score(start, span, names)
            if score > best_score:
                best_score = score
                best_start = start
                best_span = span
                best_names = names
    return best_start, best_span, best_names


def parse_uploaded_dataset(
    upload,
    manual_header_row: int | None = None,
    manual_header_span: int | None = None,
) -> ParsedDataset:
    matrix, row_numbers, source_format = _parse_matrix_from_upload(upload)
    if not matrix:
        raise ReconciliationError("No data rows found in file.")
    if len(matrix) > MAX_ROWS + HEADER_SEARCH_SPAN:
        raise ReconciliationError(
            f"File exceeds the maximum supported size of {MAX_ROWS} rows."
        )

    if manual_header_row is not None:
        header_row_index = max(0, manual_header_row - 1)
        if header_row_index >= len(matrix):
            raise ReconciliationError("Header row is outside the uploaded data.")
        header_row_span = _pad_header_rows(
            matrix, header_row_index, manual_header_span or 1
        )
        columns = _build_header_names(matrix, header_row_index, header_row_span)
        parser_notes = [
            f"Manual header configuration used: row {manual_header_row}, span {header_row_span}."
        ]
    else:
        header_row_index, header_row_span, columns = _detect_header_spec(matrix)
        parser_notes = [
            f"Auto-detected header row {header_row_index + 1} with span {header_row_span}."
        ]

    data_start = header_row_index + header_row_span
    if data_start >= len(matrix):
        raise ReconciliationError("No data rows found after header rows.")

    rows: list[dict[str, str]] = []
    normalized_rows: list[dict[str, str]] = []
    data_row_numbers: list[int] = []

    for matrix_idx in range(data_start, len(matrix)):
        row = list(matrix[matrix_idx])
        if len(row) < len(columns):
            row.extend([""] * (len(columns) - len(row)))
        else:
            row = row[: len(columns)]
        row_dict = {col: _clean_text(value) for col, value in zip(columns, row)}
        norm_dict = {
            col: _normalize_for_compare(value) for col, value in row_dict.items()
        }
        rows.append(row_dict)
        normalized_rows.append(norm_dict)
        data_row_numbers.append(
            row_numbers[matrix_idx] if matrix_idx < len(row_numbers) else matrix_idx + 1
        )

    if len(rows) > MAX_ROWS:
        raise ReconciliationError(
            f"File exceeds the maximum supported size of {MAX_ROWS} rows."
        )

    return ParsedDataset(
        rows=rows,
        normalized_rows=normalized_rows,
        columns=columns,
        column_meta=_column_meta(columns),
        header_row_index=header_row_index,
        header_row_span=header_row_span,
        parser_notes=parser_notes,
        source_format=source_format,
        row_numbers=data_row_numbers,
    )


def _parse_uploaded_dataset_fixed(
    upload, header_row: int, header_span: int, label: str
) -> ParsedDataset:
    try:
        return parse_uploaded_dataset(upload, header_row, header_span)
    except Exception as exc:
        logger.warning(
            "%s fixed header parse failed, falling back to row 1/span 1: %s", label, exc
        )
        return parse_uploaded_dataset(upload, 1, 1)


def _parse_optional_int(value: str | None) -> int | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if not re.fullmatch(r"\d+", text):
        raise ReconciliationError("Header values must be positive integers.")
    num = int(text)
    if num <= 0:
        raise ReconciliationError("Header values must be greater than 0.")
    return num


def _parse_manual_mappings(value: str | None) -> dict[str, str]:
    if not value:
        return {}
    try:
        data = json.loads(value)
    except json.JSONDecodeError:
        raise ReconciliationError("Manual mappings must be valid JSON.")
    if not isinstance(data, dict):
        raise ReconciliationError("Manual mappings must be a JSON object.")
    return {
        str(k): str(v) for k, v in data.items() if str(k).strip() and str(v).strip()
    }


def _csrf_check() -> None:
    token = request.headers.get("X-Grambook-CSRF", "")
    if token != CSRF_TOKEN:
        raise ReconciliationError(
            "Security token mismatch. Refresh the page and try again."
        )


def _session_scope_id() -> str:
    sid = session.get(CACHE_SESSION_KEY)
    if not sid:
        sid = secrets.token_urlsafe(16)
        session[CACHE_SESSION_KEY] = sid
    return sid


def _parse_page_number(value: str | None, default: int = 1) -> int:
    if value is None or not str(value).strip():
        return default
    text = str(value).strip()
    if not re.fullmatch(r"\d+", text):
        raise ReconciliationError("Page values must be positive integers.")
    page = int(text)
    if page <= 0:
        raise ReconciliationError("Page values must be greater than 0.")
    return page


def _parse_page_size(value: str | None, default: int = DEFAULT_PAGE_SIZE) -> int:
    if value is None or not str(value).strip():
        return default
    text = str(value).strip()
    if not re.fullmatch(r"\d+", text):
        raise ReconciliationError("Page size must be a positive integer.")
    page_size = int(text)
    if page_size <= 0:
        raise ReconciliationError("Page size must be greater than 0.")
    return min(page_size, MAX_PAGE_SIZE)


def _parse_date_mode(value: str | None) -> str:
    mode = _clean_text(value).casefold()
    return mode if mode in VALID_DATE_MODES else "auto"


def _request_fingerprint(
    admin_upload,
    suv_upload,
    admin_header_row: int | None,
    admin_header_span: int | None,
    suv_header_row: int | None,
    suv_header_span: int | None,
    admin_key: str | None,
    suv_key: str | None,
    date_mode: str | None = None,
    manual_mappings_raw: str | None = None,
) -> str:
    admin_bytes, admin_name = _file_bytes(admin_upload)
    suv_bytes, suv_name = _file_bytes(suv_upload)
    payload = {
        "admin": hashlib.sha256(admin_bytes).hexdigest(),
        "suv": hashlib.sha256(suv_bytes).hexdigest(),
        "admin_name": admin_name,
        "suv_name": suv_name,
        "admin_header_row": admin_header_row,
        "admin_header_span": admin_header_span,
        "suv_header_row": suv_header_row,
        "suv_header_span": suv_header_span,
        "admin_key": _clean_text(admin_key),
        "suv_key": _clean_text(suv_key),
        "date_mode": _clean_text(date_mode),
        "manual_mappings": manual_mappings_raw or "",
        "session_id": _session_scope_id(),
        "time_bucket": int(time.time() // 3600),
    }
    return hashlib.sha256(
        json.dumps(payload, sort_keys=True, default=str).encode("utf-8")
    ).hexdigest()


def _cache_result(cache_key: str, result: dict[str, Any]) -> None:
    with RESULT_CACHE_LOCK:
        RESULT_CACHE[cache_key] = result
        RESULT_CACHE_ORDER[:] = [k for k in RESULT_CACHE_ORDER if k != cache_key]
        RESULT_CACHE_ORDER.append(cache_key)
        while len(RESULT_CACHE_ORDER) > RESULT_CACHE_LIMIT:
            old_key = RESULT_CACHE_ORDER.pop(0)
            RESULT_CACHE.pop(old_key, None)


def _lookup_cached_result(cache_key: str) -> dict[str, Any] | None:
    with RESULT_CACHE_LOCK:
        result = RESULT_CACHE.get(cache_key)
        if result is None:
            return None
        RESULT_CACHE_ORDER[:] = [k for k in RESULT_CACHE_ORDER if k != cache_key]
        RESULT_CACHE_ORDER.append(cache_key)
        return result


def _slice_page(
    items: list[dict[str, Any]], page: int, page_size: int
) -> tuple[list[dict[str, Any]], dict[str, int]]:
    total_rows = len(items)
    total_pages = max(1, math.ceil(total_rows / max(page_size, 1)))
    current_page = min(max(page, 1), total_pages)
    start = (current_page - 1) * page_size
    end = start + page_size
    return items[start:end], {
        "total_rows": total_rows,
        "total_pages": total_pages,
        "current_page": current_page,
        "page_size": page_size,
    }


def _timeout_deadline(seconds: int = REQUEST_TIMEOUT_SECONDS) -> float:
    return time.monotonic() + seconds


def _check_timeout(deadline: float) -> None:
    if time.monotonic() > deadline:
        raise ReconciliationError(
            "Reconciliation timed out. Please retry with smaller files or a larger server timeout."
        )


def _resolve_key_column(selected_key: str | None, dataset: ParsedDataset) -> str:
    if not selected_key or not str(selected_key).strip():
        return _auto_detect_key_column(dataset)
    selected = _clean_text(selected_key)
    for col in dataset.columns:
        if col == selected:
            return col
    selected_norm = _normalize_header_compact(selected)
    for col in dataset.columns:
        if _normalize_header_compact(col) == selected_norm:
            return col
    return _auto_detect_key_column(dataset)


def _auto_detect_key_column(dataset: ParsedDataset) -> str:
    best_col = dataset.columns[0]
    best_score = -(10**9)
    sample_rows = dataset.normalized_rows
    for col in dataset.columns:
        values = [row.get(col, "") for row in sample_rows]
        non_empty = [v for v in values if v]
        if not non_empty:
            continue
        unique_ratio = len(set(non_empty)) / max(1, len(non_empty))
        fill_ratio = len(non_empty) / max(1, len(values))
        header_norm = _normalize_header_compact(col)
        header_bonus = 0
        if any(
            token in header_norm
            for token in ("id", "code", "ref", "no", "num", "key", "ક્રમ", "નં", "કોડ")
        ):
            header_bonus += 20
        if _guess_group(col) == "Identifier":
            header_bonus += 15
        score = (unique_ratio * 60) + (fill_ratio * 30) + header_bonus
        if score > best_score:
            best_score = score
            best_col = col
    return best_col


def _key_quality_candidates(dataset: ParsedDataset, limit: int = 3) -> list[str]:
    scored: list[tuple[float, str]] = []
    sample_rows = dataset.normalized_rows
    for col in dataset.columns:
        values = [row.get(col, "") for row in sample_rows]
        non_empty = [v for v in values if v]
        if not non_empty:
            continue
        unique_ratio = len(set(non_empty)) / max(1, len(non_empty))
        fill_ratio = len(non_empty) / max(1, len(values))
        header_norm = _normalize_header_compact(col)
        group_bonus = 15.0 if _guess_group(col) == "Identifier" else 0.0
        token_bonus = 0.0
        if any(
            token in header_norm
            for token in ("id", "code", "ref", "no", "num", "key", "ક્રમ", "નં", "કોડ")
        ):
            token_bonus += 15.0
        score = (unique_ratio * 65.0) + (fill_ratio * 25.0) + group_bonus + token_bonus
        scored.append((score, col))
    scored.sort(key=lambda item: (-item[0], item[1]))
    return [col for _, col in scored[:limit]]


def _key_quality_metrics(dataset: ParsedDataset, key_col: str) -> dict[str, Any]:
    values = [row.get(key_col, "") for row in dataset.normalized_rows]
    non_empty = [v for v in values if v]
    total = len(values)
    non_empty_count = len(non_empty)
    unique_non_empty = len(set(non_empty))
    duplicate_count = max(0, non_empty_count - unique_non_empty)
    return {
        "column": key_col,
        "total_rows": total,
        "non_empty_rows": non_empty_count,
        "missing_rows": total - non_empty_count,
        "non_empty_ratio": round(non_empty_count / max(total, 1), 4),
        "unique_ratio": round(unique_non_empty / max(non_empty_count, 1), 4),
        "duplicate_count": duplicate_count,
    }


def _validate_key_quality(
    dataset: ParsedDataset, key_col: str, label: str
) -> dict[str, Any]:
    metrics = _key_quality_metrics(dataset, key_col)
    metrics["warning"] = False
    if (
        metrics["non_empty_ratio"] < KEY_NON_EMPTY_HARD_MIN
        or metrics["unique_ratio"] < KEY_UNIQUE_HARD_MIN
    ):
        suggestions = _key_quality_candidates(dataset)
        suggestion_text = (
            ", ".join(suggestions) if suggestions else "no strong fallback key found"
        )
        metrics["warning"] = True
        metrics["warning_message"] = (
            f"{label} key column '{key_col}' is weak (non-empty {metrics['non_empty_ratio']:.0%}, "
            f"unique {metrics['unique_ratio']:.0%}). Suggested keys: {suggestion_text}."
        )
        logger.warning(metrics["warning_message"])
        return metrics
    if (
        metrics["non_empty_ratio"] < KEY_BORDERLINE_MIN
        or metrics["unique_ratio"] < 0.85
    ):
        metrics["warning"] = True
        metrics["warning_message"] = (
            f"{label} key '{key_col}' is borderline (non-empty {metrics['non_empty_ratio']:.0%}, "
            f"unique {metrics['unique_ratio']:.0%})."
        )
    return metrics


def _fuzzy_ratio(left: str, right: str) -> int:
    if not left or not right:
        return 0
    if fuzz is not None:
        return max(
            fuzz.ratio(left, right),
            fuzz.partial_ratio(left, right),
            fuzz.token_sort_ratio(left, right),
        )
    return int(SequenceMatcher(None, left, right).ratio() * 100)


def _column_pair_score(
    admin_col: str,
    suv_col: str,
    admin_norm_map: dict[str, str],
    suv_norm_map: dict[str, str],
) -> float:
    a_norm = admin_norm_map[admin_col]
    s_norm = suv_norm_map[suv_col]
    a_tokens = set(_tokenize_for_match(admin_col))
    s_tokens = set(_tokenize_for_match(suv_col))
    a_group = _guess_group(admin_col)
    s_group = _guess_group(suv_col)

    score = _fuzzy_ratio(a_norm, s_norm) * 0.5
    if a_norm == s_norm:
        score = 100.0
    if a_norm and s_norm and (a_norm.startswith(s_norm) or s_norm.startswith(a_norm)):
        score += 10.0
    shared = len(a_tokens & s_tokens)
    union = len(a_tokens | s_tokens)
    if shared:
        score += min(shared, 4) * 8.0
        if union:
            score += (shared / union) * 15.0
    if a_group != "Other" and a_group == s_group:
        score += 10.0
    elif a_group != "Other" or s_group != "Other":
        score -= 2.0
    if any(
        token in a_norm or token in s_norm
        for token in ("gs", "id", "no", "num", "code", "ref", "key", "ક્રમ", "નં")
    ):
        score += 5.0
    if "-" in a_norm or "-" in s_norm or "/" in a_norm or "/" in s_norm:
        score += 1.5
    return min(score, 100.0)


def _build_column_mapping(
    admin: ParsedDataset,
    suv: ParsedDataset,
    manual_mappings: dict[str, str] | None = None,
) -> tuple[
    list[dict[str, Any]], dict[str, str], list[str], list[str], list[dict[str, Any]]
]:
    manual_mappings = manual_mappings or {}
    admin_used: set[str] = set()
    suv_used: set[str] = set()
    mapping: dict[str, str] = {}
    pairs: list[dict[str, Any]] = []
    conflicts: list[dict[str, Any]] = []

    admin_norm_map = {col: _normalize_header_compact(col) for col in admin.columns}
    suv_norm_map = {col: _normalize_header_compact(col) for col in suv.columns}

    def _use_pair(a: str, s: str, confidence: float, source: str) -> None:
        if a in admin_used or s in suv_used:
            return
        admin_used.add(a)
        suv_used.add(s)
        mapping[a] = s
        pairs.append(
            {
                "admin_col": a,
                "suv_col": s,
                "confidence": round(float(confidence), 4),
                "source": source,
            }
        )

    def _candidate_rows(admin_col: str) -> list[tuple[float, int, int, str]]:
        admin_index = {col: idx for idx, col in enumerate(admin.columns)}
        suv_index = {col: idx for idx, col in enumerate(suv.columns)}
        remaining_suv = [c for c in suv.columns if c not in suv_used]
        scored: list[tuple[float, int, int, str]] = []
        for s in remaining_suv:
            score = _column_pair_score(admin_col, s, admin_norm_map, suv_norm_map)
            if score >= 60:
                scored.append((score, admin_index[admin_col], suv_index[s], s))
        scored.sort(key=lambda item: (-item[0], item[1], item[2], item[3]))
        return scored

    for admin_col, suv_col in manual_mappings.items():
        a = next(
            (
                col
                for col in admin.columns
                if col == admin_col
                or _normalize_header_compact(col)
                == _normalize_header_compact(admin_col)
            ),
            None,
        )
        s = next(
            (
                col
                for col in suv.columns
                if col == suv_col
                or _normalize_header_compact(col) == _normalize_header_compact(suv_col)
            ),
            None,
        )
        if a and s:
            _use_pair(a, s, 100.0, "manual")

    suv_by_norm: dict[str, list[str]] = defaultdict(list)
    for col, norm in suv_norm_map.items():
        suv_by_norm[norm].append(col)

    for admin_col in admin.columns:
        if admin_col in admin_used:
            continue
        exacts = [
            c
            for c in suv_by_norm.get(admin_norm_map[admin_col], [])
            if c not in suv_used
        ]
        if exacts:
            _use_pair(admin_col, exacts[0], 100.0, "exact")

    auto_threshold = 60.0
    for admin_col in admin.columns:
        if admin_col in admin_used:
            continue
        candidates = _candidate_rows(admin_col)
        if not candidates:
            continue
        best_score, _, _, best_suv = candidates[0]
        if best_score >= auto_threshold:
            _use_pair(admin_col, best_suv, best_score, "fuzzy")

    unmapped_admin = [c for c in admin.columns if c not in admin_used]
    unmapped_suv = [c for c in suv.columns if c not in suv_used]
    logger.info("Column map: %s", mapping)
    return pairs, mapping, unmapped_admin, unmapped_suv, []


def _build_records(dataset: ParsedDataset, key_col: str) -> list[RowRecord]:
    records: list[RowRecord] = []
    for idx, row in enumerate(dataset.rows):
        norm_row = dataset.normalized_rows[idx]
        key_display = row.get(key_col, "")
        key_norm = _normalize_key_value(key_display)
        key_missing = key_norm == ""
        fingerprint = _fingerprint_row(
            norm_row, [c for c in dataset.columns if c != key_col]
        )
        records.append(
            RowRecord(
                index=idx,
                row_number=dataset.row_numbers[idx]
                if idx < len(dataset.row_numbers)
                else idx + 1,
                key_display=key_display,
                key_norm=key_norm,
                row=row,
                norm_row=norm_row,
                fingerprint=fingerprint,
                key_missing=key_missing,
            )
        )
    return records


def _duplicate_summary(records: list[RowRecord]) -> dict[str, dict[str, int]]:
    summary: dict[str, dict[str, int]] = {}
    grouped: dict[str, dict[str, list[RowRecord]]] = defaultdict(
        lambda: defaultdict(list)
    )
    for rec in records:
        if rec.key_missing:
            continue
        grouped[rec.key_norm][rec.fingerprint].append(rec)
    for key_norm, fingerprint_groups in grouped.items():
        total = sum(len(bucket) for bucket in fingerprint_groups.values())
        unique = len(fingerprint_groups)
        summary[key_norm] = {
            "total_rows": total,
            "unique_fingerprints": unique,
            "duplicate_count": max(0, total - unique),
        }
    return summary


def _mismatch_signature(item: dict[str, Any]) -> str:
    payload = {
        "key_norm": item.get("key_norm", ""),
        "diff_cols": item.get("diff_cols", []),
        "diffs": {
            col: {
                "admin": item.get("diffs", {}).get(col, {}).get("admin", ""),
                "suvidha": item.get("diffs", {}).get(col, {}).get("suvidha", ""),
                "type": item.get("diffs", {}).get(col, {}).get("type", ""),
            }
            for col in sorted(item.get("diffs", {}).keys())
        },
    }
    return hashlib.sha256(
        json.dumps(payload, sort_keys=True, default=str).encode("utf-8")
    ).hexdigest()


def _group_mismatches(discrepancies: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    order: list[str] = []
    for item in discrepancies:
        signature = _mismatch_signature(item)
        group = grouped.get(signature)
        if group is None:
            group = {
                **item,
                "group_id": signature,
                "count": 0,
                "occurrences": [],
            }
            grouped[signature] = group
            order.append(signature)
        group["count"] += 1
        group["occurrences"].append(
            {
                "admin_row_number": item.get("admin_row_number"),
                "suvidha_row_number": item.get("suvidha_row_number"),
                "key": item.get("key"),
            }
        )
    return [grouped[sig] for sig in order]


def _normalize_key_value(value: Any) -> str:
    text = _normalize_for_compare(value)
    if not text:
        return ""
    # FIXED: preserve separators so 12-34, 1/234, and 1234 remain distinct keys.
    return SPACE_RE.sub(" ", text).casefold().strip()


def _pair_groups(
    admin_records: list[RowRecord],
    suv_records: list[RowRecord],
) -> tuple[list[tuple[RowRecord, RowRecord]], list[RowRecord], list[RowRecord]]:
    grouped_admin: dict[str, list[RowRecord]] = defaultdict(list)
    grouped_suv: dict[str, list[RowRecord]] = defaultdict(list)
    paired: list[tuple[RowRecord, RowRecord]] = []
    only_admin: list[RowRecord] = []
    only_suv: list[RowRecord] = []

    for rec in admin_records:
        if rec.key_missing:
            only_admin.append(rec)
            continue
        grouped_admin[rec.key_norm].append(rec)
    for rec in suv_records:
        if rec.key_missing:
            only_suv.append(rec)
            continue
        grouped_suv[rec.key_norm].append(rec)

    for key in sorted(set(grouped_admin) | set(grouped_suv)):
        a_list = sorted(
            grouped_admin.get(key, []), key=lambda r: (r.row_number, r.index)
        )
        s_list = sorted(grouped_suv.get(key, []), key=lambda r: (r.row_number, r.index))
        exact_admin: dict[str, list[RowRecord]] = defaultdict(list)
        exact_suv: dict[str, list[RowRecord]] = defaultdict(list)
        for rec in a_list:
            exact_admin[rec.fingerprint].append(rec)
        for rec in s_list:
            exact_suv[rec.fingerprint].append(rec)

        matched_admin_ids: set[int] = set()
        matched_suv_ids: set[int] = set()
        for fingerprint in sorted(set(exact_admin) & set(exact_suv)):
            admin_bucket = exact_admin[fingerprint]
            suv_bucket = exact_suv[fingerprint]
            pair_count = min(len(admin_bucket), len(suv_bucket))
            for i in range(pair_count):
                a_rec = admin_bucket[i]
                s_rec = suv_bucket[i]
                paired.append((a_rec, s_rec))
                matched_admin_ids.add(id(a_rec))
                matched_suv_ids.add(id(s_rec))

        remaining_admin = [rec for rec in a_list if id(rec) not in matched_admin_ids]
        remaining_suv = [rec for rec in s_list if id(rec) not in matched_suv_ids]

        pair_count = min(len(remaining_admin), len(remaining_suv))
        for i in range(pair_count):
            paired.append((remaining_admin[i], remaining_suv[i]))
        if len(remaining_admin) > pair_count:
            only_admin.extend(remaining_admin[pair_count:])
        if len(remaining_suv) > pair_count:
            only_suv.extend(remaining_suv[pair_count:])

    paired.sort(key=lambda p: (p[0].key_norm, p[0].row_number, p[1].row_number))
    only_admin.sort(key=lambda r: (r.row_number, r.index))
    only_suv.sort(key=lambda r: (r.row_number, r.index))
    return paired, only_admin, only_suv


def _compare_pair(
    admin_rec: RowRecord,
    suv_rec: RowRecord,
    admin: ParsedDataset,
    suv: ParsedDataset,
    column_map: dict[str, str],
    unmapped_suv: list[str],
    date_mode: str = "auto",
) -> dict[str, Any] | None:
    diffs: dict[str, dict[str, Any]] = {}
    diff_cols: list[str] = []

    for admin_col in admin.columns:
        suv_col = column_map.get(admin_col)
        admin_val = admin_rec.row.get(admin_col, "")
        if suv_col is None:
            diffs[admin_col] = {
                "admin": admin_val,
                "suvidha": VALUE_MISSING,
                "type": "missing_column",
            }
            diff_cols.append(admin_col)
            continue
        suv_val = suv_rec.row.get(suv_col, "")
        if not _values_equal(admin_val, suv_val, date_mode):
            diffs[admin_col] = {"admin": admin_val, "suvidha": suv_val, "type": "value"}
            diff_cols.append(admin_col)

    for suv_col in unmapped_suv:
        key = f"[SUV] {suv_col}"
        diffs[key] = {
            "admin": VALUE_MISSING,
            "suvidha": suv_rec.row.get(suv_col, ""),
            "type": "extra_column",
        }

    if not diff_cols and not unmapped_suv:
        return None

    return {
        "key": suv_rec.key_display or admin_rec.key_display,
        "key_norm": admin_rec.key_norm or suv_rec.key_norm,
        "admin_row_number": admin_rec.row_number,
        "suvidha_row_number": suv_rec.row_number,
        "admin_row": admin_rec.row,
        "suvidha": suv_rec.row,
        "suv_row": suv_rec.row,
        "diffs": diffs,
        "diff_cols": diff_cols,
        "suv_diff_cols": [f"SUV::{col}" for col in unmapped_suv],
        "key_missing": admin_rec.key_missing or suv_rec.key_missing,
    }


def reconcile(
    admin: ParsedDataset,
    suv: ParsedDataset,
    admin_key: str | None,
    suv_key: str | None,
    date_mode: str = "auto",
    manual_mappings: dict[str, str] | None = None,
) -> dict[str, Any]:
    deadline = _timeout_deadline()
    date_mode = _parse_date_mode(date_mode)
    admin_key_col = _resolve_key_column(admin_key, admin)
    suv_key_col = _resolve_key_column(suv_key, suv)
    logger.info("Using key columns: admin=%s suv=%s", admin_key_col, suv_key_col)

    admin_key_quality = _validate_key_quality(admin, admin_key_col, "Admin")
    suv_key_quality = _validate_key_quality(suv, suv_key_col, "Suvidha")
    if admin_key_quality.get("warning"):
        logger.warning(admin_key_quality.get("warning_message"))
    if suv_key_quality.get("warning"):
        logger.warning(suv_key_quality.get("warning_message"))

    col_pairs, column_map, unmapped_admin, unmapped_suv, mapping_conflicts = (
        _build_column_mapping(admin, suv, manual_mappings)
    )
    _check_timeout(deadline)
    logger.info("Unmapped admin columns: %s", unmapped_admin)
    logger.info("Unmapped suv columns: %s", unmapped_suv)

    admin_records = _build_records(admin, admin_key_col)
    suv_records = _build_records(suv, suv_key_col)
    admin_duplicate_summary = _duplicate_summary(admin_records)
    suv_duplicate_summary = _duplicate_summary(suv_records)
    admin_missing_keys = sum(1 for rec in admin_records if rec.key_missing)
    suv_missing_keys = sum(1 for rec in suv_records if rec.key_missing)
    paired, only_admin_records, only_suv_records = _pair_groups(
        admin_records, suv_records
    )

    discrepancies: list[dict[str, Any]] = []
    matched_pairs = 0
    mismatched_pairs = 0

    for admin_rec, suv_rec in paired:
        _check_timeout(deadline)
        row_diff = _compare_pair(
            admin_rec, suv_rec, admin, suv, column_map, unmapped_suv, date_mode
        )
        if row_diff is None:
            matched_pairs += 1
            continue
        mismatched_pairs += 1
        row_diff["mismatch_signature"] = _mismatch_signature(row_diff)
        discrepancies.append(row_diff)
        logger.debug(
            "Mismatch key=%s admin_row=%s suv_row=%s diff_cols=%s",
            row_diff["key"],
            row_diff["admin_row_number"],
            row_diff["suvidha_row_number"],
            row_diff["diff_cols"],
        )

    only_admin_rows = [
        {
            **record.row,
            "__row_number": record.row_number,
            "__key": record.key_display,
            "__key_norm": record.key_norm,
            "__key_missing": record.key_missing,
        }
        for record in only_admin_records
    ]
    only_suv_rows = [
        {
            **record.row,
            "__row_number": record.row_number,
            "__key": record.key_display,
            "__key_norm": record.key_norm,
            "__key_missing": record.key_missing,
        }
        for record in only_suv_records
    ]

    grouped_discrepancies = _group_mismatches(discrepancies)
    mapped_ratio = len(col_pairs) / max(min(len(admin.columns), len(suv.columns)), 1)
    weighted_total = 0.0
    weighted_weight = 0.0
    core_mapped = False
    for pair in col_pairs:
        group = _guess_group(pair["admin_col"])
        if group == "Identifier":
            weight = 3.0
            if pair["confidence"] >= MAPPING_CORE_CONFIDENCE:
                core_mapped = True
        elif group in {"Amount", "Date"}:
            weight = 2.0
        else:
            weight = 1.0
        weighted_total += pair["confidence"] * weight
        weighted_weight += weight
    weighted_confidence = weighted_total / weighted_weight if weighted_weight else 0.0

    total_groups = len(paired) + len(only_admin_records) + len(only_suv_records)
    total_rows = len(admin_records) + len(suv_records)
    logger.info(
        json.dumps(
            {
                "rows_processed": total_rows,
                "matches": matched_pairs,
                "mismatches": mismatched_pairs,
                "only_admin": len(only_admin_records),
                "only_suv": len(only_suv_records),
                "time_taken_sec": round(
                    time.monotonic() - (deadline - REQUEST_TIMEOUT_SECONDS), 4
                ),
            }
        )
    )
    logger.info(
        "Comparison complete: total=%s matched=%s mismatched=%s only_admin=%s only_suv=%s",
        total_groups,
        matched_pairs,
        mismatched_pairs,
        len(only_admin_records),
        len(only_suv_records),
    )

    return {
        "discrepancies": grouped_discrepancies,
        "only_admin_rows": only_admin_rows,
        "only_suv_rows": only_suv_rows,
        "column_map": column_map,
        "col_pairs": col_pairs,
        "mapping_conflicts": mapping_conflicts,
        "admin_key": admin_key_col,
        "suv_key": suv_key_col,
        "admin_cols": admin.columns,
        "suv_cols": suv.columns,
        "admin_col_meta": admin.column_meta,
        "suv_col_meta": suv.column_meta,
        "unmapped": {"admin_cols": unmapped_admin, "suv_cols": unmapped_suv},
        "duplicate_counts": {
            "admin": admin_duplicate_summary,
            "suv": suv_duplicate_summary,
        },
        "stats": {
            "total": total_rows,
            "groups": total_groups,
            "matched": matched_pairs,
            "mismatched": mismatched_pairs,
            "disc": mismatched_pairs,
            "only_a": len(only_admin_records),
            "only_s": len(only_suv_records),
            "compared": matched_pairs + mismatched_pairs,
            "missing_admin_keys": admin_missing_keys,
            "missing_suv_keys": suv_missing_keys,
        },
        "meta": {
            "admin_key_auto": admin_key_col == _auto_detect_key_column(admin),
            "suv_key_auto": suv_key_col == _auto_detect_key_column(suv),
            "fuzzy_threshold": FUZZY_MATCH_THRESHOLD,
            "zero_loss_warning": bool(unmapped_admin or unmapped_suv),
            "manual_mapping_count": sum(
                1 for p in col_pairs if p.get("source") == "manual"
            ),
            "date_mode": date_mode,
            "key_quality": {
                "admin": admin_key_quality,
                "suvidha": suv_key_quality,
            },
            "missing_key_counts": {
                "admin": admin_missing_keys,
                "suvidha": suv_missing_keys,
            },
            "mapping_quality": {
                "mapped_ratio": round(mapped_ratio, 4),
                "weighted_confidence": round(weighted_confidence, 4),
                "core_mapped": core_mapped,
            },
        },
    }


def _excel_styles() -> dict[str, Any]:
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        "border": border,
        "header_fill": PatternFill("solid", start_color="1F4E78"),
        "admin_fill": PatternFill("solid", start_color="FFF4F0"),
        "suv_fill": PatternFill("solid", start_color="F1FBF6"),
        "mismatch_fill_admin": PatternFill("solid", start_color="FDE2DB"),
        "mismatch_fill_suv": PatternFill("solid", start_color="DDF3E7"),
        "only_admin_fill": PatternFill("solid", start_color="FDEDED"),
        "only_suv_fill": PatternFill("solid", start_color="E6F7F5"),
        "sep_fill": PatternFill("solid", start_color="EFF2F7"),
        "header_font": Font(bold=True, color="FFFFFF", size=10),
        "bold_font": Font(bold=True, color="1F2937", size=10),
        "admin_font": Font(bold=True, color="B5451B", size=10),
        "suv_font": Font(bold=True, color="1A6B4A", size=10),
    }


def _style_cell(
    cell, *, fill=None, font=None, align="left", wrap=False, border=None
) -> None:
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border is not None:
        cell.border = border


def generate_discrepancy_report(
    admin: ParsedDataset,
    suv: ParsedDataset,
    result: dict[str, Any],
) -> tempfile.SpooledTemporaryFile:
    styles = _excel_styles()
    wb = Workbook(write_only=True)

    def _wo_cell(
        ws, value, *, fill=None, font=None, align="left", wrap=False, border=None
    ):
        cell = WriteOnlyCell(ws, value=value)
        if fill is not None:
            cell.fill = fill
        if font is not None:
            cell.font = font
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if border is not None:
            cell.border = border
        return cell

    ws_summary = wb.create_sheet("Summary")
    summary_rows = [
        ("Grambook Reconciliation Report", ""),
        ("Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Admin Key", result.get("admin_key", "")),
        ("Suvidha Key", result.get("suv_key", "")),
        ("Total Rows", result["stats"]["total"]),
        ("Total Groups", result["stats"].get("groups", "")),
        ("Matched", result["stats"]["matched"]),
        ("Mismatched", result["stats"]["mismatched"]),
        ("Only in Admin", result["stats"]["only_a"]),
        ("Only in Suvidha", result["stats"]["only_s"]),
        ("Missing Admin Keys", result["stats"].get("missing_admin_keys", "")),
        ("Missing Suvidha Keys", result["stats"].get("missing_suv_keys", "")),
        (
            "Mapping Confidence",
            result.get("meta", {})
            .get("mapping_quality", {})
            .get("weighted_confidence", ""),
        ),
    ]
    ws_summary.append(
        [
            _wo_cell(
                ws_summary,
                summary_rows[0][0],
                fill=styles["header_fill"],
                font=styles["header_font"],
                border=styles["border"],
            ),
            _wo_cell(
                ws_summary,
                summary_rows[0][1],
                fill=styles["header_fill"],
                font=styles["header_font"],
                border=styles["border"],
            ),
        ]
    )
    for label, value in summary_rows[1:]:
        ws_summary.append(
            [
                _wo_cell(
                    ws_summary,
                    label,
                    font=styles["bold_font"],
                    fill=styles["admin_fill"],
                    border=styles["border"],
                ),
                _wo_cell(ws_summary, value, border=styles["border"]),
            ]
        )

    admin_cols = admin.columns
    suv_cols = suv.columns
    suv_lookup = {
        pair["admin_col"]: pair["suv_col"] for pair in result.get("col_pairs", [])
    }

    ws_disc = wb.create_sheet("Discrepancies")
    report_columns = ["Source", "Key", "Row #"] + admin_cols
    extra_suv_cols = [col for col in suv_cols if col not in suv_lookup.values()]
    report_columns.extend([f"SUV::{col}" for col in extra_suv_cols])

    ws_disc.append(
        [
            _wo_cell(
                ws_disc,
                col,
                fill=styles["header_fill"],
                font=styles["header_font"],
                align="center",
                wrap=True,
                border=styles["border"],
            )
            for col in report_columns
        ]
    )
    for item in result.get("discrepancies", []):
        admin_row = item.get("admin_row", {})
        suv_row = item.get("suvidha", {})
        diffs = item.get("diffs", {})
        diff_cols = set(item.get("diff_cols", []))
        count = int(item.get("count", 1) or 1)
        source_label = "Admin" if count <= 1 else f"Admin x{count}"
        suv_label = "Suvidha" if count <= 1 else f"Suvidha x{count}"

        admin_cells = [
            _wo_cell(
                ws_disc,
                source_label,
                fill=styles["admin_fill"],
                font=styles["admin_font"],
                border=styles["border"],
            ),
            _wo_cell(
                ws_disc,
                item.get("key", ""),
                fill=styles["admin_fill"],
                border=styles["border"],
            ),
            _wo_cell(
                ws_disc,
                item.get("admin_row_number", ""),
                fill=styles["admin_fill"],
                border=styles["border"],
            ),
        ]
        for col in admin_cols:
            value = admin_row.get(col, "")
            suv_col = suv_lookup.get(col)
            if col in diff_cols or suv_col is None:
                admin_cells.append(
                    _wo_cell(
                        ws_disc,
                        value,
                        fill=styles["mismatch_fill_admin"],
                        font=styles["bold_font"],
                        border=styles["border"],
                        wrap=True,
                    )
                )
            else:
                admin_cells.append(
                    _wo_cell(
                        ws_disc,
                        value,
                        fill=styles["admin_fill"],
                        border=styles["border"],
                        wrap=True,
                    )
                )
        for col in extra_suv_cols:
            label = f"[SUV] {col}"
            value = diffs.get(label, {}).get("suvidha", suv_row.get(col, ""))
            admin_cells.append(
                _wo_cell(
                    ws_disc,
                    "VALUE_MISSING" if value == VALUE_MISSING else value,
                    fill=styles["mismatch_fill_admin"],
                    font=styles["bold_font"],
                    border=styles["border"],
                    wrap=True,
                )
            )
        ws_disc.append(admin_cells)

        suv_cells = [
            _wo_cell(
                ws_disc,
                suv_label,
                fill=styles["suv_fill"],
                font=styles["suv_font"],
                border=styles["border"],
            ),
            _wo_cell(
                ws_disc,
                item.get("key", ""),
                fill=styles["suv_fill"],
                border=styles["border"],
            ),
            _wo_cell(
                ws_disc,
                item.get("suvidha_row_number", ""),
                fill=styles["suv_fill"],
                border=styles["border"],
            ),
        ]
        for col in admin_cols:
            suv_col = suv_lookup.get(col)
            value = suv_row.get(suv_col, "") if suv_col else VALUE_MISSING
            if value == VALUE_MISSING:
                value = "VALUE_MISSING"
            if col in diff_cols or suv_col is None:
                suv_cells.append(
                    _wo_cell(
                        ws_disc,
                        value,
                        fill=styles["mismatch_fill_suv"],
                        font=styles["bold_font"],
                        border=styles["border"],
                        wrap=True,
                    )
                )
            else:
                suv_cells.append(
                    _wo_cell(
                        ws_disc,
                        value,
                        fill=styles["suv_fill"],
                        border=styles["border"],
                        wrap=True,
                    )
                )
        for col in extra_suv_cols:
            suv_cells.append(
                _wo_cell(
                    ws_disc,
                    suv_row.get(col, ""),
                    fill=styles["suv_fill"],
                    border=styles["border"],
                    wrap=True,
                )
            )
        ws_disc.append(suv_cells)

        ws_disc.append(
            [
                _wo_cell(ws_disc, "", fill=styles["sep_fill"], border=styles["border"])
                for _ in range(len(report_columns))
            ]
        )

    ws_admin = wb.create_sheet("Only in Admin")
    ws_admin.append(
        [
            _wo_cell(
                ws_admin,
                col,
                fill=styles["header_fill"],
                font=styles["header_font"],
                align="center",
                wrap=True,
                border=styles["border"],
            )
            for col in admin_cols
        ]
    )
    for row in result.get("only_admin_rows", []):
        ws_admin.append(
            [
                _wo_cell(
                    ws_admin,
                    row.get(col, ""),
                    fill=styles["only_admin_fill"],
                    border=styles["border"],
                    wrap=True,
                )
                for col in admin_cols
            ]
        )

    ws_suv = wb.create_sheet("Only in Suvidha")
    ws_suv.append(
        [
            _wo_cell(
                ws_suv,
                col,
                fill=styles["header_fill"],
                font=styles["header_font"],
                align="center",
                wrap=True,
                border=styles["border"],
            )
            for col in suv_cols
        ]
    )
    for row in result.get("only_suv_rows", []):
        ws_suv.append(
            [
                _wo_cell(
                    ws_suv,
                    row.get(col, ""),
                    fill=styles["only_suv_fill"],
                    border=styles["border"],
                    wrap=True,
                )
                for col in suv_cols
            ]
        )

    buf = tempfile.SpooledTemporaryFile(max_size=10 * 1024 * 1024, mode="w+b")
    wb.save(buf)
    buf.seek(0)
    return buf


def _response_items(result: dict[str, Any], view: str) -> list[dict[str, Any]]:
    if view == "oa":
        return list(result.get("only_admin_rows", []))
    if view == "os":
        return list(result.get("only_suv_rows", []))
    return list(result.get("discrepancies", []))


def _json_response(
    result: dict[str, Any], *, view: str, page: int, page_size: int
) -> dict[str, Any]:
    items = _response_items(result, view)
    page_items, pagination = _slice_page(items, page, page_size)
    return {
        "view": view,
        "data": page_items,
        "pagination": pagination,
        "stats": result.get("stats", {}),
        "matching_records": result.get("stats", {}).get("matched", 0),
        "mismatched_records": result.get("stats", {}).get("mismatched", 0),
        "col_pairs": result.get("col_pairs", []),
        "mapping_conflicts": result.get("mapping_conflicts", []),
        "column_map": result.get("column_map", {}),
        "unmapped": result.get("unmapped", {}),
        "duplicate_counts": result.get("duplicate_counts", {}),
        "admin_cols": result.get("admin_cols", []),
        "suv_cols": result.get("suv_cols", []),
        "admin_col_meta": result.get("admin_col_meta", []),
        "suv_col_meta": result.get("suv_col_meta", []),
        "meta": result.get("meta", {}),
        "only_suvidha_rows": page_items if view == "os" else [],
    }


@app.route("/")
def index():
    return send_from_directory(STATIC_DIR, "index.html")


@app.route("/api/csrf", methods=["GET"])
def api_csrf():
    return jsonify({"token": CSRF_TOKEN})


@app.route("/api/columns", methods=["POST"])
def api_columns():
    _csrf_check()
    admin_upload = request.files.get("admin_file")
    suv_upload = request.files.get("suvidha_file")
    if not admin_upload or not suv_upload:
        return jsonify({"error": "Both files are required."}), 400
    try:
        admin = _parse_uploaded_dataset_fixed(
            admin_upload, FIXED_ADMIN_HEADER_ROW, FIXED_ADMIN_HEADER_SPAN, "Admin"
        )
        suv = _parse_uploaded_dataset_fixed(
            suv_upload, FIXED_SUV_HEADER_ROW, FIXED_SUV_HEADER_SPAN, "Suvidha"
        )
        _, _, _, _, mapping_conflicts = _build_column_mapping(admin, suv, {})
        return jsonify(
            {
                "admin_cols": admin.columns,
                "admin_col_meta": admin.column_meta,
                "suv_cols": suv.columns,
                "suv_col_meta": suv.column_meta,
                "preview": {
                    "admin": {
                        "detected_header_row": admin.header_row_index + 1,
                        "detected_header_span": admin.header_row_span,
                        "notes": admin.parser_notes,
                    },
                    "suvidha": {
                        "detected_header_row": suv.header_row_index + 1,
                        "detected_header_span": suv.header_row_span,
                        "notes": suv.parser_notes,
                    },
                },
                "suggested_keys": {
                    "admin": _auto_detect_key_column(admin),
                    "suvidha": _auto_detect_key_column(suv),
                },
                "mapping_conflicts": mapping_conflicts,
            }
        )
    except ReconciliationError as exc:
        logger.warning("Reconcile rejected: %s", exc)
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:  # pragma: no cover
        logger.exception("Failed to parse columns")
        return jsonify({"error": f"Failed to parse files: {exc}"}), 500


@app.route("/api/header-preview", methods=["POST"])
def api_header_preview():
    _csrf_check()
    upload = request.files.get("file")
    if not upload:
        return jsonify({"error": "File is required."}), 400
    try:
        role = _clean_text(
            request.form.get("role") or request.args.get("role")
        ).casefold()
        if role == "suvidha":
            parsed = _parse_uploaded_dataset_fixed(
                upload, FIXED_SUV_HEADER_ROW, FIXED_SUV_HEADER_SPAN, "Preview Suvidha"
            )
        else:
            parsed = _parse_uploaded_dataset_fixed(
                upload, FIXED_ADMIN_HEADER_ROW, FIXED_ADMIN_HEADER_SPAN, "Preview Admin"
            )
        return jsonify(
            {
                "columns": parsed.columns,
                "col_meta": parsed.column_meta,
                "header_row": parsed.header_row_index + 1,
                "header_span": parsed.header_row_span,
                "sample_rows": parsed.rows[:PREVIEW_ROWS],
                "notes": parsed.parser_notes,
                "source_format": parsed.source_format,
            }
        )
    except ReconciliationError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:  # pragma: no cover
        logger.exception("Preview failed")
        return jsonify({"error": f"Preview failed: {exc}"}), 500


@app.route("/reconcile", methods=["POST"])
@app.route("/api/reconcile", methods=["POST"])
def api_reconcile():
    try:
        _csrf_check()
        admin_upload = request.files.get("admin_file")
        suv_upload = request.files.get("suvidha_file")
        if not admin_upload or not suv_upload:
            return jsonify({"error": "Both files are required."}), 400
        view = _clean_text(
            request.form.get("view") or request.args.get("view") or "disc"
        ).casefold()
        if view not in {"disc", "oa", "os"}:
            return jsonify({"error": "Invalid view."}), 400
        page = _parse_page_number(request.form.get("page") or request.args.get("page"))
        page_size = _parse_page_size(
            request.form.get("page_size") or request.args.get("page_size")
        )
        date_mode = "auto"
        cache_key = _request_fingerprint(
            admin_upload,
            suv_upload,
            FIXED_ADMIN_HEADER_ROW,
            FIXED_ADMIN_HEADER_SPAN,
            FIXED_SUV_HEADER_ROW,
            FIXED_SUV_HEADER_SPAN,
            request.form.get("admin_key"),
            request.form.get("suv_key") or request.form.get("suvidha_key"),
            date_mode,
            request.form.get("manual_mappings"),
        )
        admin = _parse_uploaded_dataset_fixed(
            admin_upload, FIXED_ADMIN_HEADER_ROW, FIXED_ADMIN_HEADER_SPAN, "Admin"
        )
        suv = _parse_uploaded_dataset_fixed(
            suv_upload, FIXED_SUV_HEADER_ROW, FIXED_SUV_HEADER_SPAN, "Suvidha"
        )
        result = reconcile(
            admin,
            suv,
            request.form.get("admin_key"),
            request.form.get("suv_key") or request.form.get("suvidha_key"),
            date_mode,
            _parse_manual_mappings(request.form.get("manual_mappings")),
        )
        _cache_result(cache_key, result)
        return jsonify(
            _json_response(result, view=view, page=page, page_size=page_size)
        )
    except ReconciliationError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:  # pragma: no cover
        logger.exception("Reconcile failed")
        return jsonify({"error": f"Reconciliation failed: {exc}"}), 500


@app.route("/api/download", methods=["POST"])
def api_download():
    try:
        _csrf_check()
        admin_upload = request.files.get("admin_file")
        suv_upload = request.files.get("suvidha_file")
        if not admin_upload or not suv_upload:
            return jsonify({"error": "Both files are required."}), 400
        date_mode = "auto"
        cache_key = _request_fingerprint(
            admin_upload,
            suv_upload,
            FIXED_ADMIN_HEADER_ROW,
            FIXED_ADMIN_HEADER_SPAN,
            FIXED_SUV_HEADER_ROW,
            FIXED_SUV_HEADER_SPAN,
            request.form.get("admin_key"),
            request.form.get("suv_key") or request.form.get("suvidha_key"),
            date_mode,
            request.form.get("manual_mappings"),
        )
        admin = _parse_uploaded_dataset_fixed(
            admin_upload, FIXED_ADMIN_HEADER_ROW, FIXED_ADMIN_HEADER_SPAN, "Admin"
        )
        suv = _parse_uploaded_dataset_fixed(
            suv_upload, FIXED_SUV_HEADER_ROW, FIXED_SUV_HEADER_SPAN, "Suvidha"
        )
        result = _lookup_cached_result(cache_key)
        if result is None:
            result = reconcile(
                admin,
                suv,
                request.form.get("admin_key"),
                request.form.get("suv_key") or request.form.get("suvidha_key"),
                date_mode,
                _parse_manual_mappings(request.form.get("manual_mappings")),
            )
            _cache_result(cache_key, result)
        buf = generate_discrepancy_report(admin, suv, result)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return send_file(
            buf,
            as_attachment=True,
            download_name=f"grambook_reconciliation_{timestamp}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except ReconciliationError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:  # pragma: no cover
        logger.exception("Download failed")
        return jsonify({"error": f"Failed to generate Excel report: {exc}"}), 500


@app.errorhandler(413)
def request_too_large(_exc):
    return jsonify(
        {"error": "Uploaded file is too large. Maximum allowed size is 50 MB."}
    ), 413


if __name__ == "__main__":
    print("Grambook Reconciliation Tool")
    print(f"Static dir: {STATIC_DIR}")
    print(f"MAX_ROWS: {MAX_ROWS}")
    app.run(debug=True, host="127.0.0.1", port=5000)
